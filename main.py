
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.proxy import Proxy, ProxyType
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.keys import Keys  # Import Keys here


def register_account(driver):
    # Navigate to the target webpage
    driver.get('https://www.google.com')
    time.sleep(2)

    driver.get('https://newconfig.nocaptchaai.com/?APIKEY=_YOUR_API_&PLANTYPE=pro&customEndpoint=&hCaptchaEnabled=true&reCaptchaEnabled=true&dataDomeEnabled=true&ocrEnabled=true&extensionEnabled=true&logsEnabled=false&fastAnimationMode=true&debugMode=false&hCaptchaAutoOpen=true&hCaptchaAutoSolve=true&hCaptchaAlwaysSolve=true&englishLanguage=true&hCaptchaGridSolveTime=7&hCaptchaMultiSolveTime=5&hCaptchaBoundingBoxSolveTime=5&reCaptchaAutoOpen=true&reCaptchaAutoSolve=true&reCaptchaAlwaysSolve=true&reCaptchaClickDelay=400&reCaptchaSubmitDelay=1&reCaptchaSolveType=image')
    
    time.sleep(3)  # Adjust the sleep time if necessary
    driver.get('chrome://extensions/')
    time.sleep(2)  # Adjust the sleep time if necessary
    driver.get("https://www.fakemail.net/")
    wait = WebDriverWait(driver, 20)
    wait.until(lambda d: d.execute_script("return document.readyState === 'complete';"))

    email_xpath1 = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="email"]')))  # Replace with your XPath
    email_value = email_xpath1.text

    # Print the text value
    print('Text value of the element:', email_value)
    time.sleep(5)

    # Open a new tab
    driver.execute_script("window.open('');")
    driver.switch_to.window(driver.window_handles[1])

    driver.get("https://elevenlabs.io/app/sign-up")
    
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, 'email'))).send_keys(email_value)
    print("email entered in Field ")
    time.sleep(2)  # Adjust the sleep time if necessary

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, 'password'))).send_keys("SecurePassword123!")
    time.sleep(2)  # Adjust the sleep time if necessary

    actions = ActionChains(driver)

    # Perform the keyboard actions: Tab, Space, Tab, Space
    actions.send_keys(Keys.TAB)  # Tab key
    actions.send_keys(Keys.SPACE)  # Space key
    actions.send_keys(Keys.TAB)  # Tab key
    actions.send_keys(Keys.SPACE)  # Space key

    actions.perform()

    # Solve the captcha
    # captcha_solution = solve_captcha(driver)

    # Attempt to click the "Sign Up" button
    time.sleep(10)  # Wait for 5 seconds before each request
    print("First Sign up button click ")
    submitbtn = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div/div[2]/div/div/form/div[6]/button")
    driver.execute_script("arguments[0].click();", submitbtn)
  
    time.sleep(20)

    driver.execute_script("arguments[0].click();", submitbtn)
    time.sleep(20)

    print("Mail Verification start")
    driver.switch_to.window(driver.window_handles[0])

    driver.refresh()

    wait = WebDriverWait(driver, 20)
    wait.until(lambda d: d.execute_script("return document.readyState === 'complete';"))
    print("Mail Verification end")
    time.sleep(5)
    driver.get("https://www.fakemail.net/window/id/2")
    time.sleep(8)
    driver.switch_to.frame("iframeMail")  # Replace with the iframe's ID

    verify_email_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Verify Email']"))
        )
    verify_email_button.click()
    time.sleep(3)
    driver.close()

    driver.switch_to.window(driver.window_handles[1])
    wait.until(lambda d: d.execute_script("return document.readyState === 'complete';"))
    time.sleep(5)

    # Re-login flow
    driver.get("https://elevenlabs.io/app/sign-in")
    time.sleep(5)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, 'email'))).send_keys(email_value)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, 'password'))).send_keys("SecurePassword123!")
    
    time.sleep(15)

    loginbtn = driver.find_element(By.XPATH, "//*[@id='sign-in-form']/div[3]/button")
    driver.execute_script("arguments[0].click();", loginbtn)
    time.sleep(10)


    if driver.current_url == "https://elevenlabs.io/app/onboarding":
        try:
            # Find the 'Skip' button using XPath
            skipbtn = driver.find_element(By.XPATH, "//button[normalize-space()='Skip']")
            
            # Execute JavaScript to click the button
            driver.execute_script("arguments[0].click();", skipbtn)
            
            # Wait for a few seconds to observe the result
            time.sleep(3)
        except Exception as e:
            print(f"An error occurred: {e}")

    time.sleep(5)
    profilebtn = driver.find_element(By.XPATH, "//p[@class='text-sm text-dark font-medium truncate inter']")
    driver.execute_script("arguments[0].click();", profilebtn)
    time.sleep(3)

    apikeybtn = driver.find_element(By.XPATH, "//span[normalize-space()='API Keys']")
    driver.execute_script("arguments[0].click();", apikeybtn)
    time.sleep(3)

    createkeybtn = driver.find_element(By.XPATH, "//button[normalize-space()='Create Key']")
    driver.execute_script("arguments[0].click();", createkeybtn)
    time.sleep(3)

    createbtnpress = driver.find_element(By.XPATH, "//button[normalize-space()='Create']")
    driver.execute_script("arguments[0].click();", createbtnpress)
    time.sleep(3)

    apikettext = driver.find_element(By.XPATH, "//p[@class='text-xs text-dark font-normal pt-2']")
    driver.execute_script("arguments[0].click();", apikettext)
    time.sleep(3)
    input_element = driver.find_element(By.XPATH, "//input[@readonly]")  # Adjust XPath as needed

    # Get the value of the read-only input field
    api_value = input_element.get_attribute("value")

    print("Read-Only Input Field Value:", api_value)
    print("excel part start ")

    time.sleep(2)
    workbook = Workbook()
    sheet = workbook.active
    data = []

    # Write data to cells   
    sheet['A1'] = 'Name'
    sheet['B1'] = 'Email'
    sheet['C1'] = 'API'
    sheet['A2'] = email_value
    sheet['B2'] = "SecurePassword123!"  # Replace with your actual password if needed
    sheet['C2'] = api_value
    # Save the Workbook



    try:
        # Try to load an existing workbook
        workbook = load_workbook('example.xlsx')
    except FileNotFoundError:
        # If the file does not exist, create a new workbook
        workbook = Workbook()
    sheet_name='Sheet1'
    # Select the sheet or create a new one if it doesn't exist
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)
    data.append([email_value, "SecurePassword123!", api_value])

    # Write data to the next available rows
    start_row = sheet.max_row + 1
    for row_index, row_data in enumerate(data, start=start_row):
        for col_index, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=row_index, column=col_index, value=cell_value)

    # Save the workbook
    workbook.save("example.xlsx")
    print("excel part end")

    driver.quit()




def main():
    # Setup WebDriver options
        options = Options()
    
        extension_path = "C:\\pythonproject\\nocaptchaai\\"
        chrome_options = Options()
        chrome_options.add_argument(f'--load-extension={extension_path}')

        # Set up ChromeDriver
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
        driver.maximize_window()
        
        # Run the automation for a specified number of rounds
    
        register_account(driver)
    
        driver.quit()

if __name__ == "__main__":
    main()
